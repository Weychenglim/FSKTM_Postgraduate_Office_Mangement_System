from datetime import date, timedelta
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZIP_STORED, ZipFile

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APITestCase

from accounts.models import OfficeStaff, Student
from academics.models import AcademicSemester
from openpyxl import Workbook, load_workbook

from dashboard.upload_security import (
    MAX_TIMELINE_ARCHIVE_ENTRIES,
    MAX_TIMELINE_UNCOMPRESSED_BYTES,
    MAX_TIMELINE_UPLOAD_BYTES,
)


User = get_user_model()


REQUIRED_HEADERS = [
    "Level",
    "Title",
    "Detail",
    "Action",
    "Deadline Start",
    "Deadline End",
    "Week Label",
    "Target Roles",
]


def workbook_upload(rows, headers=None, filename="timeline.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Timeline"
    sheet.append(headers or REQUIRED_HEADERS)
    for row in rows:
        sheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return SimpleUploadedFile(
        filename,
        stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


MINIMUM_XLSX_ENTRIES = {
    "[Content_Types].xml": b"content-types-data",
    "_rels/.rels": b"relationships-data",
    "xl/workbook.xml": b"workbook-data",
    "xl/worksheets/sheet1.xml": b"worksheet-data",
}


def archive_upload(entries, filename="timeline.xlsx", compression=ZIP_STORED):
    stream = BytesIO()
    with ZipFile(stream, "w", compression=compression) as archive:
        for name, content in entries.items():
            archive.writestr(name, content)
    return SimpleUploadedFile(
        filename,
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def encrypted_flag_upload():
    upload = archive_upload(MINIMUM_XLSX_ENTRIES)
    data = bytearray(upload.read())
    local_header = data.find(b"PK\x03\x04")
    central_header = data.find(b"PK\x01\x02")
    for offset in (local_header + 6, central_header + 8):
        flags = int.from_bytes(data[offset : offset + 2], "little") | 1
        data[offset : offset + 2] = flags.to_bytes(2, "little")
    return SimpleUploadedFile("encrypted.xlsx", bytes(data))


class DashboardTimelineApiTests(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            email="office@example.com",
            password="password123",
            full_name="Office Staff",
            role=User.Role.OFFICE_ADMIN,
        )
        OfficeStaff.objects.create(
            user=self.admin,
            staff_no="A10001",
            department="Postgraduate Office Division",
        )
        self.student = User.objects.create_user(
            email="student-timeline@example.com",
            password="password123",
            full_name="Timeline Student",
            role=User.Role.STUDENT,
        )
        Student.objects.create(user=self.student, matric_no="S10001", programme="MASTER OF ARTIFICIAL INTELLIGENCE (COURSEWORK)")
        today = timezone.localdate()
        self.academic_semester = AcademicSemester.objects.create(
            code=f"{today.year}-{today.year + 1}-S1",
            academic_session=f"{today.year}/{today.year + 1}",
            term=AcademicSemester.Term.SEMESTER_I,
            starts_on=today - timedelta(days=30),
            ends_on=today + timedelta(days=120),
            lifecycle_status=AcademicSemester.Lifecycle.ACTIVE,
            created_by=self.admin,
            activated_at=timezone.now(),
        )

    def authenticate(self, user):
        self.client.force_authenticate(user=user)

    def valid_rows(self):
        return [
            [
                "P1",
                "Supervisor appointment request",
                "Students submit appointment of supervisor forms.",
                "Student / Supervisor",
                date(2026, 3, 16),
                date(2026, 3, 20),
                "Week 2",
                "STUDENT,LECTURER",
            ],
            [
                "P1",
                "Step 2 decision notice",
                "Office informs students and supervisors of step 2 decision.",
                "TDIT Office",
                date(2026, 4, 3),
                date(2026, 4, 3),
                "",
                "OFFICE_STAFF",
            ],
            [
                "P2",
                "Final presentation",
                "Students complete the final presentation.",
                "Student / Supervisor / Examiner",
                date(2026, 6, 8),
                date(2026, 7, 3),
                "Week 13 - 15",
                "STUDENT,LECTURER",
            ],
        ]

    def test_active_timeline_returns_not_available_payload_when_no_active_timeline_exists(self):
        self.authenticate(self.student)

        response = self.client.get("/api/dashboard/timeline/active/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["available"], False)
        self.assertEqual(response.data["message"], "No timeline available at now")
        self.assertEqual(response.data["levels"], [])

    def test_only_office_admin_can_download_template_and_upload_timeline(self):
        self.authenticate(self.student)

        template_response = self.client.get("/api/dashboard/timeline/template/")
        upload_response = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": workbook_upload(self.valid_rows())},
            format="multipart",
        )

        self.assertEqual(template_response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(upload_response.status_code, status.HTTP_403_FORBIDDEN)

        self.authenticate(self.admin)
        admin_template = self.client.get("/api/dashboard/timeline/template/")

        self.assertEqual(admin_template.status_code, status.HTTP_200_OK)
        self.assertIn("spreadsheetml.sheet", admin_template["Content-Type"])
        self.assertIn("FSKTM_Semester_Timeline_Template.xlsx", admin_template["Content-Disposition"])
        workbook = load_workbook(SimpleUploadedFile("template.xlsx", admin_template.content))
        headers = [cell.value for cell in workbook.active[1]]
        self.assertEqual(headers, REQUIRED_HEADERS)
        self.assertNotIn("Step", headers)
        self.assertNotIn("Status", headers)

    def test_valid_upload_creates_active_grouped_timeline_and_audit_log(self):
        self.authenticate(self.admin)

        response = self.client.post(
            "/api/dashboard/timeline/upload/",
            {
                "semesterId": self.academic_semester.pk,
                "file": workbook_upload(self.valid_rows(), filename="sem2-timeline.xlsx"),
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["importedCount"], 3)
        self.assertEqual(response.data["timeline"]["semester"], "Semester I")
        self.assertEqual(
            response.data["timeline"]["session"],
            self.academic_semester.academic_session,
        )
        self.assertEqual(
            response.data["timeline"]["semesterId"],
            self.academic_semester.pk,
        )
        self.assertEqual(response.data["timeline"]["levels"][0]["level"], "P1")
        self.assertEqual(len(response.data["timeline"]["levels"][0]["entries"]), 2)
        self.assertEqual(response.data["timeline"]["levels"][1]["level"], "P2")

        active = self.client.get("/api/dashboard/timeline/active/")
        self.assertEqual(active.status_code, status.HTTP_200_OK)
        self.assertEqual(active.data["available"], True)
        self.assertEqual(active.data["levels"][0]["entries"][0]["step"], 1)
        self.assertEqual(active.data["levels"][0]["entries"][0]["title"], "Supervisor appointment request")
        self.assertEqual(active.data["levels"][0]["entries"][0]["detail"], "Students submit appointment of supervisor forms.")
        self.assertEqual(active.data["levels"][0]["entries"][0]["targetRoles"], ["STUDENT", "LECTURER"])

        from dashboard.models import TimelineAuditLog

        self.assertEqual(TimelineAuditLog.objects.count(), 1)
        self.assertEqual(TimelineAuditLog.objects.first().action, "UPLOAD")

    def test_invalid_upload_reports_template_errors(self):
        self.authenticate(self.admin)

        response = self.client.post(
            "/api/dashboard/timeline/upload/",
            {
                "file": workbook_upload(
                    [["P1", "Missing required fields"]],
                    headers=["Level", "Title"],
                    filename="bad-template.xlsx",
                )
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("errors", response.data)
        self.assertTrue(any("Missing required columns" in error for error in response.data["errors"]))

    def test_upload_rejects_files_larger_than_ten_megabytes(self):
        self.authenticate(self.admin)
        upload = SimpleUploadedFile(
            "oversized.xlsx",
            b"x" * (MAX_TIMELINE_UPLOAD_BYTES + 1),
        )

        response = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(response.data, {"errors": ["Timeline file must be 10 MB or smaller."]})

    def test_upload_rejects_renamed_non_xlsx_extension(self):
        self.authenticate(self.admin)

        response = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": workbook_upload(self.valid_rows(), filename="timeline.zip")},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            response.data,
            {"errors": ["Only .xlsx timeline templates are accepted."]},
        )

    def test_upload_rejects_malformed_zip_container(self):
        self.authenticate(self.admin)

        response = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": SimpleUploadedFile("malformed.xlsx", b"not-a-zip")},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            response.data,
            {"errors": ["Timeline file is not a valid XLSX workbook."]},
        )

    def test_upload_rejects_missing_xlsx_workbook_structure(self):
        self.authenticate(self.admin)

        response = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": archive_upload({"notes.txt": b"not a workbook"})},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            response.data,
            {"errors": ["Timeline file is missing required XLSX workbook structure."]},
        )

    def test_upload_rejects_corrupt_zip_entry(self):
        self.authenticate(self.admin)
        upload = archive_upload(MINIMUM_XLSX_ENTRIES)
        data = bytearray(upload.read())
        payload_offset = data.find(b"workbook-data")
        data[payload_offset] ^= 0xFF

        response = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": SimpleUploadedFile("corrupt.xlsx", bytes(data))},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            response.data,
            {"errors": ["Timeline XLSX archive failed its integrity check."]},
        )

    def test_upload_rejects_encrypted_archive_entries(self):
        self.authenticate(self.admin)

        response = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": encrypted_flag_upload()},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            response.data,
            {"errors": ["Encrypted XLSX archive entries are not allowed."]},
        )

    def test_upload_rejects_archive_path_traversal(self):
        self.authenticate(self.admin)
        entries = {**MINIMUM_XLSX_ENTRIES, "../outside.txt": b"blocked"}

        response = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": archive_upload(entries)},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            response.data,
            {"errors": ["XLSX archive contains an unsafe file path."]},
        )

    def test_upload_rejects_embedded_macro_payload(self):
        self.authenticate(self.admin)
        entries = {**MINIMUM_XLSX_ENTRIES, "xl/vbaProject.bin": b"macro"}

        response = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": archive_upload(entries)},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            response.data,
            {"errors": ["Macro-enabled XLSX content is not allowed."]},
        )

    def test_upload_rejects_more_than_one_thousand_archive_entries(self):
        self.authenticate(self.admin)
        entries = dict(MINIMUM_XLSX_ENTRIES)
        for index in range(MAX_TIMELINE_ARCHIVE_ENTRIES - len(entries) + 1):
            entries[f"xl/media/item-{index}.txt"] = b"x"

        response = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": archive_upload(entries, compression=ZIP_DEFLATED)},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            response.data,
            {"errors": ["XLSX archive contains more than 1,000 entries."]},
        )

    def test_upload_rejects_more_than_fifty_megabytes_uncompressed(self):
        self.authenticate(self.admin)
        entries = {
            **MINIMUM_XLSX_ENTRIES,
            "xl/media/expanded.txt": b"0" * (MAX_TIMELINE_UNCOMPRESSED_BYTES + 1),
        }

        response = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": archive_upload(entries, compression=ZIP_DEFLATED)},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(
            response.data,
            {"errors": ["XLSX archive expands beyond the 50 MB limit."]},
        )

    def test_upload_rejects_invalid_date_range(self):
        self.authenticate(self.admin)

        response = self.client.post(
            "/api/dashboard/timeline/upload/",
            {
                "file": workbook_upload(
                    [
                        [
                            "P1",
                            "Invalid range",
                            "First row",
                            "Student",
                            date(2026, 3, 20),
                            date(2026, 3, 16),
                            "Week 2",
                            "STUDENT",
                        ],
                    ],
                    filename="conflicting-template.xlsx",
                )
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertTrue(any("Deadline End cannot be before Deadline Start" in error for error in response.data["errors"]))

    def test_upload_rejects_target_roles_outside_student_lecturer_office_staff(self):
        self.authenticate(self.admin)

        response = self.client.post(
            "/api/dashboard/timeline/upload/",
            {
                "file": workbook_upload(
                    [
                        [
                            "P1",
                            "All role event",
                            "Should not accept ALL as a target role.",
                            "Student",
                            date(2026, 3, 16),
                            date(2026, 3, 20),
                            "Week 2",
                            "ALL",
                        ],
                    ],
                    filename="invalid-role-template.xlsx",
                )
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertTrue(any("Invalid target role(s): ALL" in error for error in response.data["errors"]))

    def test_upload_replaces_previous_active_timeline(self):
        self.authenticate(self.admin)
        first = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"semester": "Semester I", "session": "2025/2026", "file": workbook_upload(self.valid_rows())},
            format="multipart",
        )
        self.assertEqual(first.status_code, status.HTTP_201_CREATED)

        second = self.client.post(
            "/api/dashboard/timeline/upload/",
            {
                "semester": "Semester II",
                "session": "2025/2026",
                "file": workbook_upload(
                    [["P2", "Marks entry", "Student marks entry period.", "Student / Examiner", date(2026, 6, 8), date(2026, 7, 10), "Week 13 - 16", "STUDENT,LECTURER"]],
                    filename="replacement.xlsx",
                ),
            },
            format="multipart",
        )

        self.assertEqual(second.status_code, status.HTTP_201_CREATED)

        from dashboard.models import SemesterTimeline

        self.assertEqual(SemesterTimeline.objects.count(), 2)
        self.assertEqual(SemesterTimeline.objects.filter(is_active=True).count(), 1)
        self.assertEqual(SemesterTimeline.objects.get(is_active=True).semester, "Semester I")

    def test_office_admin_can_patch_timeline_entry_and_audit_change(self):
        self.authenticate(self.admin)
        upload = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": workbook_upload(self.valid_rows())},
            format="multipart",
        )
        entry_id = upload.data["timeline"]["levels"][0]["entries"][0]["id"]

        today = date.today()
        response = self.client.patch(
            f"/api/dashboard/timeline/entries/{entry_id}/",
            {
                "deadlineStart": (today - timedelta(days=1)).isoformat(),
                "deadlineEnd": (today + timedelta(days=1)).isoformat(),
                "weekLabel": "Week 2 - Updated",
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["status"], "Active")
        self.assertEqual(response.data["weekLabel"], "Week 2 - Updated")

        from dashboard.models import TimelineAuditLog

        self.assertTrue(TimelineAuditLog.objects.filter(action="EDIT_ENTRY").exists())

    def test_office_admin_can_move_timeline_entry_between_project_levels(self):
        self.authenticate(self.admin)
        upload = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": workbook_upload(self.valid_rows())},
            format="multipart",
        )
        entry_id = upload.data["timeline"]["levels"][0]["entries"][0]["id"]

        response = self.client.patch(
            f"/api/dashboard/timeline/entries/{entry_id}/",
            {"level": "P2"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["level"], "P2")
        self.assertEqual(response.data["step"], 2)

        active = self.client.get("/api/dashboard/timeline/active/")
        p2_entries = [group for group in active.data["levels"] if group["level"] == "P2"][0]["entries"]
        self.assertTrue(any(entry["id"] == entry_id for entry in p2_entries))

    def test_office_admin_can_create_timeline_entry_and_audit_change(self):
        self.authenticate(self.admin)
        upload = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": workbook_upload(self.valid_rows())},
            format="multipart",
        )
        self.assertEqual(upload.status_code, status.HTTP_201_CREATED)

        today = date.today()
        upcoming_start = today + timedelta(days=7)
        upcoming_end = today + timedelta(days=14)
        response = self.client.post(
            "/api/dashboard/timeline/entries/",
            {
                "level": "P1",
                "title": "Proposal presentation",
                "detail": "Proposal presentation",
                "action": "Student / Supervisor / Examiner",
                "deadlineStart": upcoming_start.isoformat(),
                "deadlineEnd": upcoming_end.isoformat(),
                "weekLabel": "Week 13 - 17",
                "targetRoles": ["STUDENT", "LECTURER"],
                "status": "Completed",
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

        response = self.client.post(
            "/api/dashboard/timeline/entries/",
            {
                "level": "P1",
                "title": "Proposal presentation",
                "detail": "Proposal presentation",
                "action": "Student / Supervisor / Examiner",
                "deadlineStart": upcoming_start.isoformat(),
                "deadlineEnd": upcoming_end.isoformat(),
                "weekLabel": "Week 13 - 17",
                "targetRoles": ["STUDENT", "LECTURER"],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["level"], "P1")
        self.assertEqual(response.data["title"], "Proposal presentation")
        self.assertEqual(response.data["step"], 3)
        self.assertEqual(response.data["status"], "Upcoming")

        active = self.client.get("/api/dashboard/timeline/active/")
        p1_entries = active.data["levels"][0]["entries"]
        self.assertEqual(len(p1_entries), 3)
        self.assertEqual(p1_entries[-1]["title"], "Proposal presentation")

        from dashboard.models import TimelineAuditLog

        self.assertTrue(TimelineAuditLog.objects.filter(action="ADD_ENTRY").exists())

    def test_office_admin_can_delete_timeline_entry_and_audit_change(self):
        self.authenticate(self.admin)
        upload = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": workbook_upload(self.valid_rows())},
            format="multipart",
        )
        entry_id = upload.data["timeline"]["levels"][0]["entries"][0]["id"]

        response = self.client.delete(f"/api/dashboard/timeline/entries/{entry_id}/")

        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)

        active = self.client.get("/api/dashboard/timeline/active/")
        remaining_ids = [
            entry["id"]
            for group in active.data["levels"]
            for entry in group["entries"]
        ]
        self.assertNotIn(entry_id, remaining_ids)

        from dashboard.models import TimelineAuditLog

        self.assertTrue(TimelineAuditLog.objects.filter(action="DELETE_ENTRY").exists())

    def test_office_admin_can_view_real_timeline_audit_logs(self):
        self.authenticate(self.admin)
        upload = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": workbook_upload(self.valid_rows(), filename="timeline-audit.xlsx")},
            format="multipart",
        )
        entry_id = upload.data["timeline"]["levels"][0]["entries"][0]["id"]
        self.client.patch(
            f"/api/dashboard/timeline/entries/{entry_id}/",
            {"weekLabel": "Week 2 - Updated"},
            format="json",
        )

        response = self.client.get("/api/dashboard/timeline/audit-logs/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["logs"][0]["action"], "EDIT_ENTRY")
        self.assertEqual(response.data["logs"][0]["actorName"], "Office Staff")
        self.assertTrue(any(log["action"] == "UPLOAD" for log in response.data["logs"]))

    def test_non_admin_cannot_create_delete_or_view_timeline_audit_logs(self):
        self.authenticate(self.admin)
        upload = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": workbook_upload(self.valid_rows())},
            format="multipart",
        )
        entry_id = upload.data["timeline"]["levels"][0]["entries"][0]["id"]

        self.authenticate(self.student)

        create_response = self.client.post(
            "/api/dashboard/timeline/entries/",
            {
                "level": "P1",
                "title": "Blocked",
                "detail": "Blocked",
                "action": "Student",
                "deadlineStart": "2026-06-08",
                "deadlineEnd": "2026-06-08",
                "targetRoles": ["STUDENT"],
            },
            format="json",
        )
        delete_response = self.client.delete(f"/api/dashboard/timeline/entries/{entry_id}/")
        audit_response = self.client.get("/api/dashboard/timeline/audit-logs/")

        self.assertEqual(create_response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(delete_response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(audit_response.status_code, status.HTTP_403_FORBIDDEN)

    def test_office_staff_tasks_include_timeline_owner_tasks(self):
        from dashboard.models import SemesterTimelineEntry

        self.authenticate(self.admin)
        upload = self.client.post(
            "/api/dashboard/timeline/upload/",
            {"file": workbook_upload(self.valid_rows())},
            format="multipart",
        )
        self.assertEqual(upload.status_code, status.HTTP_201_CREATED)
        active_entry = SemesterTimelineEntry.objects.get(
            title="Step 2 decision notice",
        )
        active_entry.deadline_start = timezone.localdate()
        active_entry.deadline_end = timezone.localdate()
        active_entry.save(update_fields=["deadline_start", "deadline_end"])

        response = self.client.get("/api/dashboard/tasks/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        task_names = [task["name"] for task in response.data["tasks"]]
        self.assertIn("Step 2 decision notice", task_names)
        self.assertIn("Upload semester timeline", task_names)

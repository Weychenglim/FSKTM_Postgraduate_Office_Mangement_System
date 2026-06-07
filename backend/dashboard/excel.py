from datetime import date, datetime
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .models import SemesterTimelineEntry


REQUIRED_HEADERS = [
    "Level",
    "Title",
    "Detail",
    "Action",
    "Deadline Start",
    "Deadline End",
    "Week Label",
    "Target Roles",
]


def normalize_header(value):
    return str(value or "").strip()


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError("Invalid date format.")


def parse_target_roles(value):
    raw = str(value or "").replace("/", ",")
    roles = [role.strip().upper().replace(" ", "_") for role in raw.split(",") if role.strip()]
    invalid = [role for role in roles if role not in SemesterTimelineEntry.VALID_TARGET_ROLES]
    if invalid:
        raise ValueError(f"Invalid target role(s): {', '.join(invalid)}")
    if not roles:
        raise ValueError("At least one target role is required.")
    return roles


def derive_status(deadline_start, deadline_end):
    today = timezone.localdate()
    if today < deadline_start:
        return SemesterTimelineEntry.Status.UPCOMING
    if today > deadline_end:
        return SemesterTimelineEntry.Status.COMPLETED
    if deadline_start == deadline_end:
        return SemesterTimelineEntry.Status.DEADLINE
    return SemesterTimelineEntry.Status.ACTIVE


def parse_timeline_workbook(uploaded_file):
    errors = []
    parsed_rows = []
    try:
        workbook = load_workbook(uploaded_file, data_only=True)
    except Exception as exc:
        return [], [f"Could not read Excel workbook: {exc}"]

    sheet = workbook.active
    headers = [normalize_header(cell.value) for cell in sheet[1]]
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        return [], [f"Missing required columns: {', '.join(missing)}"]

    column_index = {header: headers.index(header) for header in REQUIRED_HEADERS}
    next_step_by_level = {"P1": 0, "P2": 0}

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in row):
            continue

        def get(header):
            index = column_index[header]
            return row[index] if index < len(row) else None

        row_errors = []
        level = str(get("Level") or "").strip().upper()
        if level not in {"P1", "P2"}:
            row_errors.append("Level must be P1 or P2.")

        title = str(get("Title") or "").strip()
        if not title:
            row_errors.append("Title is required.")

        detail = str(get("Detail") or "").strip()
        if not detail:
            row_errors.append("Detail is required.")

        action = str(get("Action") or "").strip()
        if not action:
            row_errors.append("Action is required.")

        try:
            deadline_start = parse_date(get("Deadline Start"))
        except ValueError:
            deadline_start = None
            row_errors.append("Deadline Start must be a valid date.")

        try:
            deadline_end = parse_date(get("Deadline End"))
        except ValueError:
            deadline_end = None
            row_errors.append("Deadline End must be a valid date.")

        if deadline_start and deadline_end and deadline_end < deadline_start:
            row_errors.append("Deadline End cannot be before Deadline Start.")

        week_label = str(get("Week Label") or "").strip()

        try:
            target_roles = parse_target_roles(get("Target Roles"))
        except ValueError as exc:
            target_roles = []
            row_errors.append(str(exc))

        if row_errors:
            errors.extend([f"Row {row_number}: {error}" for error in row_errors])
            continue

        next_step_by_level[level] += 1
        parsed_rows.append(
            {
                "level": level,
                "step": next_step_by_level[level],
                "title": title,
                "detail": detail,
                "action_owner": action,
                "deadline_start": deadline_start,
                "deadline_end": deadline_end,
                "week_label": week_label,
                "target_roles": target_roles,
                "status": derive_status(deadline_start, deadline_end),
                "display_order": len(parsed_rows) + 1,
            }
        )

    if not parsed_rows and not errors:
        errors.append("Timeline workbook does not contain any timeline rows.")

    return parsed_rows, errors


def build_template_workbook():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Timeline"
    sheet.append(REQUIRED_HEADERS)
    sheet.append(
        [
            "P1",
            "Supervisor appointment request",
            "Students submit appointment of supervisor forms.",
            "Student / Supervisor",
            date(2026, 3, 16),
            date(2026, 3, 20),
            "Week 2",
            "STUDENT,LECTURER",
        ]
    )
    sheet.append(
        [
            "P2",
            "Final presentation",
            "Students complete the final presentation.",
            "Student / Supervisor / Examiner",
            date(2026, 6, 8),
            date(2026, 7, 3),
            "Week 13 - 15",
            "STUDENT,LECTURER",
        ]
    )
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 14), 48)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.getvalue()

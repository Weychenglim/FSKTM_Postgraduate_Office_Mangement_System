from collections import Counter
from datetime import date, datetime
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.exceptions import ObjectDoesNotExist
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.exceptions import PermissionDenied, ValidationError

from academics.models import AcademicSemester
from appointments.ageing import panel_waiting_metadata, supervisor_waiting_metadata
from appointments.models import PanelRecommendation, SupervisorApplication
from marks.deadlines import mark_deadline_metadata
from marks.models import EvaluationTask, MarkEntry

from .models import SemesterTimelineEntry


User = get_user_model()
AUTHORIZED_ROLES = {
    User.Role.OFFICE_ADMIN,
    User.Role.COORDINATOR,
    User.Role.LECTURER,
}
AGE_BANDS = ("0-3", "4-7", "8-14", "15+")
REPORT_ROLE_NAMES = {
    User.Role.OFFICE_ADMIN: "OFFICE_ADMIN",
    User.Role.COORDINATOR: "COORDINATOR",
    User.Role.LECTURER: "LECTURER",
}


def _coordinator_programme(user):
    try:
        return user.lecturer.coordinator.programme_managed.strip()
    except (AttributeError, ObjectDoesNotExist):
        return ""


def _parse_optional_date(value, field_name):
    if value in (None, ""):
        return None
    try:
        parsed = parse_date(str(value))
    except ValueError:
        parsed = None
    if parsed is None:
        raise ValidationError({field_name: f"{field_name} must be an ISO date."})
    return parsed


def _date_of(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return timezone.localtime(value).date() if timezone.is_aware(value) else value.date()
    if isinstance(value, date):
        return value
    return None


def _within_range(value, start_date, end_date):
    record_date = _date_of(value)
    if record_date is None:
        return start_date is None and end_date is None
    if start_date and record_date < start_date:
        return False
    if end_date and record_date > end_date:
        return False
    return True


def _iso(value):
    return value.isoformat() if value is not None else None


def _age_band(days):
    if days is None:
        return None
    if days <= 3:
        return "0-3"
    if days <= 7:
        return "4-7"
    if days <= 14:
        return "8-14"
    return "15+"


def _counts(values):
    return dict(sorted(Counter(values).items()))


def _resolve_semester_filter(params):
    selector = str(params.get("semester") or "active").strip()
    available = list(AcademicSemester.objects.order_by("-starts_on", "-id"))
    if selector in {"all", "unassigned"}:
        return selector, None, available
    if selector == "active":
        semester = next(
            (
                item
                for item in available
                if item.lifecycle_status == AcademicSemester.Lifecycle.ACTIVE
                and item.is_active
            ),
            None,
        )
        return selector, semester, available
    semester = next((item for item in available if item.code == selector), None)
    if semester is None:
        raise ValidationError(
            {"semester": "Select active, all, unassigned, or a valid semester code."}
        )
    return selector, semester, available


def _filter_semester(queryset, lookup, selector, semester):
    if selector == "all":
        return queryset
    if selector == "unassigned":
        return queryset.filter(**{f"{lookup}__isnull": True})
    if semester is None:
        return queryset.none()
    return queryset.filter(**{lookup: semester})


def _semester_row(semester):
    return {
        "semesterId": semester.pk if semester else None,
        "semesterCode": semester.code if semester else None,
        "semester": semester.label if semester else "Legacy / Unassigned",
    }


def _module_summary(records):
    waiting_records = [record for record in records if record["waitingDays"] is not None]
    return {
        "total": len(records),
        "statusCounts": _counts(record["status"] for record in records),
        "waitingOwnerCounts": _counts(
            record["waitingOn"] for record in waiting_records if record["waitingOn"]
        ),
        "ageBands": {
            band: sum(record["ageBand"] == band for record in waiting_records)
            for band in AGE_BANDS
        },
        "records": records,
    }


def _scope_supervisor_records(user, programme, selector, semester):
    records = SupervisorApplication.objects.select_related(
        "student",
        "student__user",
        "proposed_supervisor",
        "academic_semester",
    ).prefetch_related("workflow_events")
    if user.role == User.Role.COORDINATOR:
        records = (
            records.filter(student__programme=programme)
            if programme
            else records.none()
        )
    elif user.role == User.Role.LECTURER:
        records = records.filter(proposed_supervisor=user)
    elif programme:
        records = records.filter(student__programme=programme)
    return _filter_semester(
        records,
        "academic_semester",
        selector,
        semester,
    )


def _scope_panel_records(user, programme, selector, semester):
    records = PanelRecommendation.objects.select_related(
        "profile",
        "supervisor",
        "recommended_member",
        "academic_semester",
    ).prefetch_related("workflow_events")
    if user.role == User.Role.COORDINATOR:
        records = (
            records.filter(profile__programme=programme)
            if programme
            else records.none()
        )
    elif user.role == User.Role.LECTURER:
        records = records.filter(recommended_member=user)
    elif programme:
        records = records.filter(profile__programme=programme)
    return _filter_semester(
        records,
        "academic_semester",
        selector,
        semester,
    )


def _supervisor_rows(
    user,
    programme,
    start_date,
    end_date,
    now,
    selector,
    semester,
):
    rows = []
    for application in _scope_supervisor_records(
        user,
        programme,
        selector,
        semester,
    ):
        if not _within_range(application.submitted_at, start_date, end_date):
            continue
        waiting = supervisor_waiting_metadata(application, now=now)
        rows.append(
            {
                "recordType": "SUPERVISOR_APPLICATION",
                "recordId": str(application.pk),
                "studentId": application.student.matric_no,
                "studentName": application.student.user.full_name,
                "programme": application.student.programme,
                "researchTitle": application.research_title,
                "researchArea": application.research_area,
                "assignee": application.proposed_supervisor.full_name,
                "status": application.status,
                "reportDate": _iso(application.submitted_at),
                "waitingSince": _iso(waiting["waitingSince"]),
                "waitingDays": waiting["waitingDays"],
                "waitingOn": waiting["waitingOn"],
                "ageBand": _age_band(waiting["waitingDays"]),
                **_semester_row(application.academic_semester),
            }
        )
    return rows


def _panel_rows(
    user,
    programme,
    start_date,
    end_date,
    now,
    selector,
    semester,
):
    rows = []
    for recommendation in _scope_panel_records(
        user,
        programme,
        selector,
        semester,
    ):
        report_date = recommendation.submitted_at or recommendation.created_at
        if not _within_range(report_date, start_date, end_date):
            continue
        waiting = panel_waiting_metadata(recommendation, now=now)
        rows.append(
            {
                "recordType": "PANEL_RECOMMENDATION",
                "recordId": str(recommendation.pk),
                "studentId": recommendation.profile.matric_no,
                "studentName": recommendation.profile.student_name,
                "programme": recommendation.profile.programme,
                "researchTitle": recommendation.profile.proposed_topic,
                "researchArea": recommendation.profile.research_area,
                "assignee": recommendation.recommended_member.full_name,
                "status": recommendation.status,
                "reportDate": _iso(report_date),
                "waitingSince": _iso(waiting["waitingSince"]),
                "waitingDays": waiting["waitingDays"],
                "waitingOn": waiting["waitingOn"],
                "ageBand": _age_band(waiting["waitingDays"]),
                **_semester_row(recommendation.academic_semester),
            }
        )
    return rows


def _mark_rows(
    user,
    programme,
    start_date,
    end_date,
    now,
    selector,
    semester,
):
    if user.role == User.Role.COORDINATOR:
        return None
    tasks = EvaluationTask.objects.select_related(
        "profile",
        "evaluator",
        "period",
        "period__academic_semester",
        "mark_entry",
    )
    if user.role == User.Role.LECTURER:
        tasks = tasks.filter(evaluator=user)
    elif programme:
        tasks = tasks.filter(profile__programme=programme)
    tasks = _filter_semester(
        tasks,
        "period__academic_semester",
        selector,
        semester,
    )

    rows = []
    for task in tasks:
        report_date = task.period.closes_at or task.assigned_at
        if not _within_range(report_date, start_date, end_date):
            continue
        try:
            entry = task.mark_entry
        except MarkEntry.DoesNotExist:
            entry = None
        entry_status = entry.status if entry else MarkEntry.Status.NOT_STARTED
        deadline = mark_deadline_metadata(
            task.period.closes_at,
            is_submitted=entry_status == MarkEntry.Status.SUBMITTED,
            now=now,
        )
        rows.append(
            {
                "recordType": "MARK_TASK",
                "recordId": str(task.pk),
                "studentId": task.profile.matric_no,
                "studentName": task.profile.student_name,
                "programme": task.profile.programme,
                "assignee": task.evaluator.full_name,
                "status": entry_status,
                "evaluatorRole": task.evaluator_role,
                "reportDate": _iso(report_date),
                **_semester_row(task.period.academic_semester),
                **deadline,
                "dueAt": _iso(deadline["dueAt"]),
            }
        )
    return rows


def _timeline_status(entry, today):
    if today < entry.deadline_start:
        return "UPCOMING"
    if today > entry.deadline_end:
        return "COMPLETED"
    if entry.deadline_start == entry.deadline_end:
        return "DEADLINE"
    return "ACTIVE"


def _timeline_rows(
    user,
    start_date,
    end_date,
    today,
    selector,
    semester,
):
    if user.role == User.Role.COORDINATOR:
        return None
    rows = []
    entries = SemesterTimelineEntry.objects.select_related(
        "timeline",
        "timeline__academic_semester",
    ).filter(
        timeline__is_active=True,
    )
    entries = _filter_semester(
        entries,
        "timeline__academic_semester",
        selector,
        semester,
    )
    for entry in entries:
        if user.role == User.Role.LECTURER and "LECTURER" not in entry.target_roles:
            continue
        if not _within_range(entry.deadline_start, start_date, end_date):
            continue
        rows.append(
            {
                "recordType": "TIMELINE_ENTRY",
                "recordId": str(entry.pk),
                "title": entry.title or entry.detail,
                "level": entry.level,
                "status": _timeline_status(entry, today),
                "targetRoles": entry.target_roles,
                "reportDate": entry.deadline_start.isoformat(),
                "dueAt": entry.deadline_end.isoformat(),
                **_semester_row(entry.timeline.academic_semester),
            }
        )
    return rows


def _marks_summary(rows):
    if rows is None:
        return None
    submitted = sum(row["status"] == MarkEntry.Status.SUBMITTED for row in rows)
    return {
        "total": len(rows),
        "statusCounts": _counts(row["status"] for row in rows),
        "deadlineStateCounts": _counts(row["deadlineState"] for row in rows),
        "evaluatorRoleCounts": _counts(row["evaluatorRole"] for row in rows),
        "completionRate": round((submitted / len(rows)) * 100, 1) if rows else None,
        "records": rows,
    }


def _timeline_summary(rows):
    if rows is None:
        return None
    role_counts = Counter()
    for row in rows:
        role_counts.update(row["targetRoles"])
    return {
        "total": len(rows),
        "statusCounts": _counts(row["status"] for row in rows),
        "levelCounts": _counts(row["level"] for row in rows),
        "targetRoleCounts": dict(sorted(role_counts.items())),
        "records": rows,
    }


def _attention(supervisor_rows, panel_rows, mark_rows):
    items = []
    for row in [*supervisor_rows, *panel_rows]:
        if row["waitingDays"] is None:
            continue
        items.append(
            {
                "kind": "WAITING",
                "recordType": row["recordType"],
                "recordId": row["recordId"],
                "studentId": row["studentId"],
                "label": f'{row["studentName"]} - {row["status"]}',
                "programme": row["programme"],
                "waitingDays": row["waitingDays"],
                "waitingOn": row["waitingOn"],
                "deadlineState": None,
                "dueAt": None,
            }
        )
    for row in mark_rows or []:
        if row["deadlineState"] != "OVERDUE":
            continue
        items.append(
            {
                "kind": "DEADLINE",
                "recordType": row["recordType"],
                "recordId": row["recordId"],
                "studentId": row["studentId"],
                "label": f'{row["studentName"]} - {row["evaluatorRole"]} marks',
                "programme": row["programme"],
                "waitingDays": None,
                "waitingOn": None,
                "deadlineState": row["deadlineState"],
                "dueAt": row["dueAt"],
            }
        )
    return sorted(
        items,
        key=lambda item: (
            0 if item["kind"] == "DEADLINE" else 1,
            -(item["waitingDays"] or 0),
            item["label"],
        ),
    )[:20]


def build_workflow_report(user, query_params=None, now=None):
    if user.role not in AUTHORIZED_ROLES:
        raise PermissionDenied("Workflow reports are not available for this role.")

    params = query_params or {}
    start_date = _parse_optional_date(params.get("startDate"), "startDate")
    end_date = _parse_optional_date(params.get("endDate"), "endDate")
    if start_date and end_date and start_date > end_date:
        raise ValidationError({"endDate": "endDate must be on or after startDate."})

    now = now or timezone.now()
    semester_selector, selected_semester, available_semesters = (
        _resolve_semester_filter(params)
    )
    programme = None
    if user.role == User.Role.COORDINATOR:
        programme = _coordinator_programme(user)
    elif user.role == User.Role.OFFICE_ADMIN:
        programme = str(params.get("programme") or "").strip() or None

    available_programmes = sorted(
        {
            value.strip()
            for value in list(
                SupervisorApplication.objects.values_list(
                    "student__programme", flat=True
                )
            )
            + list(
                PanelRecommendation.objects.values_list(
                    "profile__programme", flat=True
                )
            )
            + list(
                EvaluationTask.objects.values_list(
                    "profile__programme", flat=True
                )
            )
            if value and value.strip()
        }
    ) if user.role == User.Role.OFFICE_ADMIN else ([programme] if programme else [])

    supervisor_rows = _supervisor_rows(
        user,
        programme,
        start_date,
        end_date,
        now,
        semester_selector,
        selected_semester,
    )
    panel_rows = _panel_rows(
        user,
        programme,
        start_date,
        end_date,
        now,
        semester_selector,
        selected_semester,
    )
    mark_rows = _mark_rows(
        user,
        programme,
        start_date,
        end_date,
        now,
        semester_selector,
        selected_semester,
    )
    timeline_rows = _timeline_rows(
        user,
        start_date,
        end_date,
        timezone.localtime(now).date(),
        semester_selector,
        selected_semester,
    )
    waiting_days = [
        row["waitingDays"]
        for row in [*supervisor_rows, *panel_rows]
        if row["waitingDays"] is not None
    ]
    all_sections = [
        supervisor_rows,
        panel_rows,
        mark_rows or [],
        timeline_rows or [],
    ]

    return {
        "generatedAt": now.isoformat(),
        "scope": {
            "role": REPORT_ROLE_NAMES[user.role],
            "programme": programme,
        },
        "filters": {
            "startDate": start_date.isoformat() if start_date else None,
            "endDate": end_date.isoformat() if end_date else None,
            "programme": programme,
            "availableProgrammes": available_programmes,
            "semester": semester_selector,
            "selectedSemester": (
                _semester_row(selected_semester)
                if selected_semester
                else None
            ),
            "availableSemesters": [
                {
                    **_semester_row(item),
                    "lifecycleStatus": item.lifecycle_status,
                    "effectiveStatus": item.effective_status,
                }
                for item in available_semesters
            ],
        },
        "overview": {
            "totalRecords": sum(len(section) for section in all_sections),
            "pendingApprovals": len(waiting_days),
            "averageWaitingDays": (
                round(sum(waiting_days) / len(waiting_days), 1)
                if waiting_days
                else None
            ),
            "longestWaitingDays": max(waiting_days) if waiting_days else None,
            "marksCompletionRate": (
                _marks_summary(mark_rows)["completionRate"]
                if mark_rows is not None
                else None
            ),
            "overdueMarks": sum(
                row["deadlineState"] == "OVERDUE" for row in mark_rows or []
            ),
            "activeTimelineEntries": sum(
                row["status"] in {"ACTIVE", "DEADLINE"}
                for row in timeline_rows or []
            ),
        },
        "supervisor": _module_summary(supervisor_rows),
        "panel": _module_summary(panel_rows),
        "marks": _marks_summary(mark_rows),
        "timeline": _timeline_summary(timeline_rows),
        "attention": _attention(supervisor_rows, panel_rows, mark_rows),
    }


def _append_sheet(workbook, title, headers, rows):
    sheet = workbook.create_sheet(title=title)
    sheet.append([label for _key, label in headers])
    for row in rows:
        sheet.append([row.get(key) for key, _label in headers])
    for column_index, (_key, label) in enumerate(headers, start=1):
        sheet.cell(1, column_index, label)
        sheet.cell(1, column_index).font = Font(bold=True, color="FFFFFF")
        sheet.cell(1, column_index).fill = PatternFill(
            "solid", fgColor="1E3A5F"
        )
        column_widths = [
            len(label),
            *[
                len(str(sheet.cell(row_index, column_index).value or ""))
                for row_index in range(2, sheet.max_row + 1)
            ],
        ]
        width = min(42, max(column_widths) + 2)
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    return sheet


def build_workflow_report_workbook(report):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary_rows = [
        ("Generated At", report["generatedAt"]),
        ("Role", report["scope"]["role"]),
        ("Programme", report["scope"]["programme"] or "All authorised programmes"),
        ("Start Date", report["filters"]["startDate"] or "All"),
        ("End Date", report["filters"]["endDate"] or "All"),
        ("Semester", report["filters"]["semester"]),
    ] + [
        (key, value) for key, value in report["overview"].items()
    ]
    for row in summary_rows:
        summary.append(row)
    summary["A1"].font = Font(bold=True)
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 30

    workflow_headers = [
        ("recordId", "Record ID"),
        ("semesterId", "Semester ID"),
        ("semesterCode", "Semester Code"),
        ("semester", "Semester"),
        ("studentId", "Student ID"),
        ("studentName", "Student Name"),
        ("programme", "Programme"),
        ("researchTitle", "Research Title"),
        ("researchArea", "Research Area"),
        ("assignee", "Assigned To"),
        ("status", "Status"),
        ("reportDate", "Report Date"),
        ("waitingSince", "Waiting Since"),
        ("waitingDays", "Waiting Days"),
        ("waitingOn", "Waiting On"),
        ("ageBand", "Age Band"),
    ]
    _append_sheet(
        workbook,
        "Supervisor",
        workflow_headers,
        report["supervisor"]["records"],
    )
    _append_sheet(
        workbook,
        "Panel",
        workflow_headers,
        report["panel"]["records"],
    )
    if report["marks"] is not None:
        _append_sheet(
            workbook,
            "Marks",
            [
                ("recordId", "Record ID"),
                ("semesterId", "Semester ID"),
                ("semesterCode", "Semester Code"),
                ("semester", "Semester"),
                ("studentId", "Student ID"),
                ("studentName", "Student Name"),
                ("programme", "Programme"),
                ("assignee", "Evaluator"),
                ("evaluatorRole", "Evaluator Role"),
                ("status", "Status"),
                ("deadlineState", "Deadline State"),
                ("dueAt", "Due At"),
                ("daysUntilDue", "Days Until Due"),
            ],
            report["marks"]["records"],
        )
    if report["timeline"] is not None:
        timeline_rows = [
            {**row, "targetRoles": ", ".join(row["targetRoles"])}
            for row in report["timeline"]["records"]
        ]
        _append_sheet(
            workbook,
            "Timeline",
            [
                ("recordId", "Record ID"),
                ("semesterId", "Semester ID"),
                ("semesterCode", "Semester Code"),
                ("semester", "Semester"),
                ("title", "Title"),
                ("level", "Level"),
                ("status", "Status"),
                ("targetRoles", "Target Roles"),
                ("reportDate", "Start Date"),
                ("dueAt", "End Date"),
            ],
            timeline_rows,
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()

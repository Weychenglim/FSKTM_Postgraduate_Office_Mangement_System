from datetime import timedelta
from io import BytesIO

from django.contrib.auth import get_user_model
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from accounts.models import (
    Coordinator,
    Lecturer,
    OfficeStaff,
    Panel,
    Student,
    Supervisor,
)
from academics.capacity_services import (
    capacity_plan_content_fingerprint,
    clone_capacity_plan,
    publish_capacity_plan,
    update_capacity_entry,
)
from academics.models import AcademicSemester, LecturerAvailabilityWindow
from academics.test_capacity_helpers import publish_test_capacity_plan
from appointments.models import (
    PanelRecommendation,
    StudentResearchProfile,
    SupervisorApplication,
)
from dashboard.models import SemesterTimeline, SemesterTimelineEntry
from marks.models import EvaluationPeriod, EvaluationTask, MarkEntry, Rubric

User = get_user_model()
PROGRAMME = "MASTER OF ARTIFICIAL INTELLIGENCE (COURSEWORK)"
FOREIGN_PROGRAMME = "MASTER OF DATA SCIENCE"


class WorkflowReportTests(APITestCase):
    def setUp(self):
        self.now = timezone.now()
        self.office = self._user(
            "report-office@example.test", "Report Office", User.Role.OFFICE_ADMIN
        )
        OfficeStaff.objects.create(
            user=self.office,
            staff_no="REPORT-OFFICE",
            department="Postgraduate Office",
        )
        today = timezone.localdate()
        self.academic_semester = AcademicSemester.objects.create(
            code=f"{today.year}-{today.year + 1}-S1",
            academic_session=f"{today.year}/{today.year + 1}",
            term=AcademicSemester.Term.SEMESTER_I,
            starts_on=today - timedelta(days=30),
            ends_on=today + timedelta(days=120),
            lifecycle_status=AcademicSemester.Lifecycle.ACTIVE,
            created_by=self.office,
            activated_at=self.now,
        )
        self.supervisor = self._lecturer(
            "report-supervisor@example.test", "Report Supervisor", "REPORT-SUP"
        )
        Supervisor.objects.create(
            lecturer=self.supervisor.lecturer,
            max_supervisees=5,
        )
        self.panel = self._lecturer(
            "report-panel@example.test", "Report Panel", "REPORT-PANEL"
        )
        Panel.objects.create(
            lecturer=self.panel.lecturer,
            max_appointments=5,
        )
        self.foreign_panel = self._lecturer(
            "report-foreign-panel@example.test",
            "Report Foreign Panel",
            "REPORT-FOREIGN",
        )
        Panel.objects.create(
            lecturer=self.foreign_panel.lecturer,
            max_appointments=5,
        )
        self.coordinator = self._user(
            "report-coordinator@example.test",
            "Report Coordinator",
            User.Role.COORDINATOR,
        )
        Lecturer.objects.create(
            user=self.coordinator,
            staff_no="REPORT-COORD",
            department="Artificial Intelligence",
        )
        Coordinator.objects.create(
            lecturer=self.coordinator.lecturer,
            programme_managed=PROGRAMME,
        )
        self.student_user, self.student, self.profile = self._student(
            "report-student@example.test",
            "Report Student",
            "REPORT-STUDENT",
            PROGRAMME,
        )
        (
            self.foreign_student_user,
            self.foreign_student,
            self.foreign_profile,
        ) = self._student(
            "report-foreign-student@example.test",
            "Report Foreign Student",
            "REPORT-FOREIGN-STUDENT",
            FOREIGN_PROGRAMME,
        )

        self.pending_supervisor = SupervisorApplication.objects.create(
            academic_semester=self.academic_semester,
            student=self.student,
            proposed_supervisor=self.supervisor,
            research_title="Owned supervisor request",
            research_abstract="Owned request",
            status=SupervisorApplication.Status.SUBMITTED_TO_SUPERVISOR,
            submitted_at=self.now - timedelta(days=5),
        )
        self.approved_supervisor = SupervisorApplication.objects.create(
            academic_semester=self.academic_semester,
            student=self.foreign_student,
            proposed_supervisor=self.supervisor,
            research_title="Foreign approved request",
            research_abstract="Foreign request",
            status=SupervisorApplication.Status.REJECTED_BY_COORDINATOR,
            submitted_at=self.now - timedelta(days=10),
            coordinator_decided_at=self.now - timedelta(days=2),
        )
        self.pending_panel = PanelRecommendation.objects.create(
            academic_semester=self.academic_semester,
            profile=self.profile,
            supervisor=self.supervisor,
            recommended_member=self.panel,
            status=PanelRecommendation.Status.SUBMITTED_TO_PANEL,
            submitted_at=self.now - timedelta(days=8),
        )
        self.foreign_panel_record = PanelRecommendation.objects.create(
            academic_semester=self.academic_semester,
            profile=self.foreign_profile,
            supervisor=self.supervisor,
            recommended_member=self.foreign_panel,
            status=PanelRecommendation.Status.REJECTED_BY_PANEL,
            submitted_at=self.now - timedelta(days=3),
            panel_decided_at=self.now - timedelta(days=1),
        )

        rubric = Rubric.objects.create(name="Report Rubric", code="report-rubric")
        self.period = EvaluationPeriod.objects.create(
            academic_semester=self.academic_semester,
            name="Report Period",
            semester="Semester 1 2026/2027",
            rubric=rubric,
            closes_at=self.now - timedelta(days=1),
            is_open=True,
        )
        self.mark_task = EvaluationTask.objects.create(
            profile=self.profile,
            evaluator=self.panel,
            period=self.period,
            evaluator_role=EvaluationTask.EvaluatorRole.PANEL,
        )
        MarkEntry.objects.create(task=self.mark_task, status=MarkEntry.Status.DRAFT)

        timeline = SemesterTimeline.objects.create(
            academic_semester=self.academic_semester,
            semester="Semester 1",
            session="2026/2027",
            is_active=True,
            uploaded_by=self.office,
        )
        self.timeline_entry = SemesterTimelineEntry.objects.create(
            timeline=timeline,
            level=SemesterTimelineEntry.Level.P1,
            step=1,
            title="Report milestone",
            detail="Reporting timeline milestone",
            action_owner="Lecturer",
            deadline_start=timezone.localdate(),
            deadline_end=timezone.localdate() + timedelta(days=2),
            target_roles=["LECTURER"],
            display_order=1,
        )
        publish_test_capacity_plan(self.academic_semester, self.office)

    def _user(self, email, name, role):
        return User.objects.create_user(
            email=email,
            password="password123",
            full_name=name,
            role=role,
        )

    def _lecturer(self, email, name, staff_no):
        user = self._user(email, name, User.Role.LECTURER)
        Lecturer.objects.create(
            user=user,
            staff_no=staff_no,
            department="Artificial Intelligence",
        )
        return user

    def _student(self, email, name, matric_no, programme):
        user = self._user(email, name, User.Role.STUDENT)
        student = Student.objects.create(
            user=user,
            matric_no=matric_no,
            programme=programme,
        )
        profile = StudentResearchProfile.objects.create(
            student=user,
            matric_no=matric_no,
            student_name=name,
            programme=programme,
            semester="Semester 1 2026/2027",
            proposed_topic=f"{name} reporting topic",
            supervisor=self.supervisor,
        )
        return user, student, profile

    def authenticate(self, user):
        self.client.force_authenticate(user=user)

    def test_office_report_aggregates_all_modules_and_programme_filter(self):
        self.authenticate(self.office)

        response = self.client.get("/api/dashboard/reports/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["scope"]["role"], "OFFICE_ADMIN")
        self.assertEqual(response.data["overview"]["totalRecords"], 6)
        self.assertEqual(response.data["overview"]["pendingApprovals"], 2)
        self.assertEqual(response.data["overview"]["longestWaitingDays"], 8)
        self.assertEqual(response.data["overview"]["overdueMarks"], 1)
        self.assertEqual(response.data["supervisor"]["ageBands"]["4-7"], 1)
        self.assertEqual(response.data["panel"]["ageBands"]["8-14"], 1)
        self.assertTrue(all(item["studentId"] for item in response.data["attention"]))
        self.assertIn(PROGRAMME, response.data["filters"]["availableProgrammes"])
        self.assertIn(
            FOREIGN_PROGRAMME, response.data["filters"]["availableProgrammes"]
        )

        marks_only_programme = "MASTER OF SOFTWARE ENGINEERING"
        _, _, marks_only_profile = self._student(
            "report-marks-only@example.test",
            "Report Marks Only Student",
            "REPORT-MARKS-ONLY",
            marks_only_programme,
        )
        EvaluationTask.objects.create(
            profile=marks_only_profile,
            evaluator=self.panel,
            period=self.period,
            evaluator_role=EvaluationTask.EvaluatorRole.PANEL,
        )
        programmes_response = self.client.get("/api/dashboard/reports/")
        self.assertIn(
            marks_only_programme,
            programmes_response.data["filters"]["availableProgrammes"],
        )

        filtered = self.client.get("/api/dashboard/reports/", {"programme": PROGRAMME})
        self.assertEqual(filtered.status_code, status.HTTP_200_OK)
        self.assertEqual(filtered.data["filters"]["programme"], PROGRAMME)
        self.assertEqual(filtered.data["supervisor"]["total"], 1)
        self.assertEqual(filtered.data["panel"]["total"], 1)

    def test_office_report_exposes_capacity_distribution_attention_and_export_columns(
        self,
    ):
        current = self.academic_semester.capacity_plans.get()
        replacement = clone_capacity_plan(current, actor=self.office)
        update_capacity_entry(
            replacement,
            lecturer=self.panel.lecturer,
            actor=self.office,
            supervisor_limit=None,
            panel_limit=0,
            expected_fingerprint=capacity_plan_content_fingerprint(replacement),
        )
        publish_capacity_plan(
            replacement,
            actor=self.office,
            reason="Exercise over-capacity reporting.",
            expected_fingerprint=capacity_plan_content_fingerprint(replacement),
        )
        today = timezone.localdate()
        LecturerAvailabilityWindow.objects.create(
            academic_semester=self.academic_semester,
            lecturer=self.supervisor.lecturer,
            role=LecturerAvailabilityWindow.Role.SUPERVISOR,
            starts_on=today,
            ends_on=today + timedelta(days=2),
            reason="Private report test reason.",
            created_by=self.office,
        )
        self.authenticate(self.office)

        response = self.client.get("/api/dashboard/reports/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response.data["capacity"]["supervisor"]["TEMPORARILY_UNAVAILABLE"], 1
        )
        self.assertEqual(response.data["capacity"]["panel"]["OVER_CAPACITY"], 1)
        capacity_attention = [
            item
            for item in response.data["attention"]
            if item["kind"] == "LECTURER_CAPACITY"
        ]
        self.assertEqual(len(capacity_attention), 2)
        self.assertTrue(
            all(
                item["recordType"] == "LECTURER_CAPACITY" for item in capacity_attention
            )
        )
        self.assertNotIn("private report", str(capacity_attention).lower())

        actions = self.client.get("/api/dashboard/tasks/")
        self.assertEqual(actions.status_code, status.HTTP_200_OK)
        capacity_actions = [
            item
            for item in actions.data["tasks"]
            if item["recordType"] == "LECTURER_CAPACITY"
        ]
        self.assertEqual(len(capacity_actions), 2)
        self.assertNotIn("private report", str(capacity_actions).lower())

        export = self.client.get("/api/dashboard/reports/export/")
        workbook = load_workbook(BytesIO(export.content))
        for sheet_name in ("Supervisor", "Panel"):
            headers = [cell.value for cell in workbook[sheet_name][1]]
            self.assertIn("Capacity Plan ID", headers)
            self.assertIn("Capacity Plan Version", headers)
            self.assertIn("Capacity State", headers)
            self.assertIn("Capacity Limit", headers)
            self.assertIn("Capacity Load", headers)
            self.assertIn("Unavailable Until", headers)

    def test_non_office_reports_do_not_expose_capacity_summary(self):
        for user in (self.coordinator, self.panel):
            self.authenticate(user)
            response = self.client.get("/api/dashboard/reports/")
            self.assertEqual(response.status_code, status.HTTP_200_OK)
            self.assertIsNone(response.data["capacity"])

    def test_report_defaults_active_and_supports_all_and_unassigned(self):
        SupervisorApplication.objects.create(
            student=self.student,
            proposed_supervisor=self.supervisor,
            research_title="Legacy unassigned request",
            research_abstract="Preserved without inferred semester.",
            status=SupervisorApplication.Status.REJECTED_BY_SUPERVISOR,
            submitted_at=self.now - timedelta(days=20),
            supervisor_decided_at=self.now - timedelta(days=19),
        )
        self.authenticate(self.office)

        active = self.client.get("/api/dashboard/reports/")
        all_semesters = self.client.get(
            "/api/dashboard/reports/",
            {"semester": "all"},
        )
        unassigned = self.client.get(
            "/api/dashboard/reports/",
            {"semester": "unassigned"},
        )

        self.assertEqual(active.data["filters"]["semester"], "active")
        self.assertEqual(active.data["supervisor"]["total"], 2)
        self.assertEqual(all_semesters.data["supervisor"]["total"], 3)
        self.assertEqual(unassigned.data["supervisor"]["total"], 1)
        self.assertEqual(
            unassigned.data["supervisor"]["records"][0]["semester"],
            "Legacy / Unassigned",
        )

    def test_coordinator_scope_cannot_be_widened_and_hides_marks(self):
        self.authenticate(self.coordinator)

        response = self.client.get(
            "/api/dashboard/reports/", {"programme": FOREIGN_PROGRAMME}
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["scope"]["programme"], PROGRAMME)
        self.assertEqual(response.data["filters"]["programme"], PROGRAMME)
        self.assertEqual(response.data["supervisor"]["total"], 1)
        self.assertEqual(response.data["panel"]["total"], 1)
        self.assertIsNone(response.data["marks"])
        self.assertIsNone(response.data["timeline"])

    def test_lecturer_report_contains_only_assigned_work(self):
        self.authenticate(self.panel)

        response = self.client.get("/api/dashboard/reports/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["supervisor"]["total"], 0)
        self.assertEqual(response.data["panel"]["total"], 1)
        self.assertEqual(response.data["marks"]["total"], 1)
        self.assertEqual(response.data["timeline"]["total"], 1)
        attention_ids = {item["recordId"] for item in response.data["attention"]}
        self.assertIn(str(self.pending_panel.pk), attention_ids)
        self.assertNotIn(str(self.foreign_panel_record.pk), attention_ids)

    def test_student_report_access_is_forbidden(self):
        self.authenticate(self.student_user)

        response = self.client.get("/api/dashboard/reports/")

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_report_validates_and_applies_date_range(self):
        self.authenticate(self.office)

        malformed = self.client.get(
            "/api/dashboard/reports/", {"startDate": "not-a-date"}
        )
        impossible = self.client.get(
            "/api/dashboard/reports/", {"startDate": "2026-02-31"}
        )
        reversed_range = self.client.get(
            "/api/dashboard/reports/",
            {"startDate": "2026-08-01", "endDate": "2026-07-01"},
        )
        empty = self.client.get(
            "/api/dashboard/reports/", {"startDate": "", "endDate": ""}
        )
        future = self.client.get(
            "/api/dashboard/reports/",
            {"startDate": "2099-01-01", "endDate": "2099-01-31"},
        )

        self.assertEqual(malformed.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(impossible.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(reversed_range.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(empty.status_code, status.HTTP_200_OK)
        self.assertEqual(future.status_code, status.HTTP_200_OK)
        self.assertEqual(future.data["overview"]["totalRecords"], 0)

    def test_export_uses_role_visible_sheets_and_frozen_headers(self):
        self.authenticate(self.coordinator)

        response = self.client.get("/api/dashboard/reports/export/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["Summary", "Supervisor", "Panel"])
        self.assertEqual(workbook["Supervisor"].freeze_panes, "A2")
        self.assertEqual(workbook["Panel"].freeze_panes, "A2")
        supervisor_ids = {str(cell.value) for cell in workbook["Supervisor"]["A"][1:]}
        self.assertIn(str(self.pending_supervisor.pk), supervisor_ids)
        self.assertNotIn(str(self.approved_supervisor.pk), supervisor_ids)

    def test_export_supports_visible_sheets_with_no_data_rows(self):
        self.authenticate(self.office)

        response = self.client.get(
            "/api/dashboard/reports/export/",
            {"startDate": "2099-01-01", "endDate": "2099-01-31"},
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(
            workbook.sheetnames,
            ["Summary", "Supervisor", "Panel", "Marks", "Timeline"],
        )
        self.assertEqual(workbook["Supervisor"].max_row, 1)
        self.assertEqual(workbook["Marks"].max_row, 1)
